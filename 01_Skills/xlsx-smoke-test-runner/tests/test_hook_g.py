"""TDD-Tests für Hook-§G Sparrate-Σ-Sanity (SPEC §4.3, Tier-Modell Umstrukturierung 2026-06-07) + P13 sheets-Expand.

Tier-Modell (löst flaches 38/19-Modell ab):
  effektive Rate = satelliten_tier_raten[tier] × DEFCON-Modulation × FLAG
    D3/D4 → voll | D2 → halb | D1 → 0 | FLAG → 0 (heilig)
  §G prüft SOLL-Invariante: Σ(tier_raten[s.tier]) == brokers.scalable.sparrate_eur (Anker 364€)
  + Funded-Echo-Display (Content-Scan 'Funded'+'SOLL', layout-robust statt fixe N19).
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

_HOOK = Path(__file__).resolve().parents[3] / "03_Tools" / "precommit" / "xlsx_smoke_test.py"

# Tier-Basisraten (== config.yaml satelliten_tier_raten)
_TR = {1: 40, 2: 32, 3: 18}


def _load_hook():
    spec = importlib.util.spec_from_file_location("hook_under_test", _HOOK)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["hook_under_test"] = mod
    spec.loader.exec_module(mod)
    return mod


# ----- _derive_rate_eur unit-tests (SPEC §4.3 Tier-Modell) -----


def test_derive_rate_eur_flag_true_returns_zero():
    mod = _load_hook()
    assert mod._derive_rate_eur({"tier": 1, "flag": True, "defcon": 3}, _TR) == 0
    assert mod._derive_rate_eur({"tier": 2, "flag": True, "defcon": 2}, _TR) == 0


def test_derive_rate_eur_flag_false_d3_d4_returns_full_tier():
    mod = _load_hook()
    assert mod._derive_rate_eur({"tier": 1, "flag": False, "defcon": 3}, _TR) == 40
    assert mod._derive_rate_eur({"tier": 2, "flag": False, "defcon": 4}, _TR) == 32
    assert mod._derive_rate_eur({"tier": 3, "flag": False, "defcon": 3}, _TR) == 18


def test_derive_rate_eur_flag_false_d2_returns_half_tier():
    mod = _load_hook()
    assert mod._derive_rate_eur({"tier": 2, "flag": False, "defcon": 2}, _TR) == 16
    assert mod._derive_rate_eur({"tier": 1, "flag": False, "defcon": 2}, _TR) == 20
    assert mod._derive_rate_eur({"tier": 3, "flag": False, "defcon": 2}, _TR) == 9


def test_derive_rate_eur_flag_false_d1_returns_zero():
    mod = _load_hook()
    assert mod._derive_rate_eur({"tier": 1, "flag": False, "defcon": 1}, _TR) == 0


def test_derive_rate_eur_unmapped_defcon_raises():
    mod = _load_hook()
    with pytest.raises(ValueError):
        mod._derive_rate_eur({"tier": 1, "flag": False, "defcon": 5}, _TR)


def test_derive_rate_eur_unmapped_tier_raises():
    mod = _load_hook()
    with pytest.raises(ValueError):
        mod._derive_rate_eur({"tier": 9, "flag": False, "defcon": 3}, _TR)


# ----- _check_g_sparrate_sigma -----


def _build_sat_wb_clean(funded: int = 210, soll: int = 364) -> openpyxl.Workbook:
    """Minimaler Sat-konformer Workbook: 2 Sheets + Funded-Echo + K3-Display (Tier-Modell)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Satelliten Monitor"
    ws["K3"] = "Ergebnis: ● 8 Volle Tier-Rate  ● 1 D2-Sockelbetrag (V)  ● 4 Eingefroren"
    ws["N20"] = f"→ Funded {funded} € (SOLL {soll} €)"
    wb.create_sheet("QuickScreen Ampel")
    return wb


def _write_config_yaml(
    tmp_path: Path, satelliten: list[dict], anker: int = 364, tier_raten: dict | None = None
) -> Path:
    p = tmp_path / "config.yaml"
    cfg = {
        "brokers": {"scalable": {"sparrate_eur": anker}},
        "satelliten_tier_raten": tier_raten if tier_raten is not None else dict(_TR),
        "satelliten": satelliten,
    }
    p.write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    return p


def _live_config_yaml_path() -> Path:
    return _HOOK.resolve().parents[2] / "01_Skills" / "dynastie-depot" / "config.yaml"


def _build_canonical_sat_list_364() -> list[dict]:
    """13-Roster Tier-Modell: SOLL-Σ = 4×40 + 3×32 + 6×18 = 364, Funded-Σ = 210 (empirie Umstrukturierung 2026-06-07)."""
    return [
        {"ticker": "ASML", "tier": 2, "defcon": 3, "flag": False},
        {"ticker": "AVGO", "tier": 1, "defcon": 2, "flag": True},
        {"ticker": "MSFT", "tier": 1, "defcon": 2, "flag": True},
        {"ticker": "RMS", "tier": 3, "defcon": 3, "flag": False},
        {"ticker": "SU", "tier": 3, "defcon": 3, "flag": False},
        {"ticker": "BRK.B", "tier": 3, "defcon": 3, "flag": False},
        {"ticker": "V", "tier": 2, "defcon": 2, "flag": False},
        {"ticker": "TMO", "tier": 3, "defcon": 3, "flag": False},
        {"ticker": "APH", "tier": 3, "defcon": 2, "flag": True},
        {"ticker": "AMZN", "tier": 1, "defcon": 1, "flag": True},
        {"ticker": "NOW", "tier": 1, "defcon": 3, "flag": False},
        {"ticker": "KYCCF", "tier": 2, "defcon": 3, "flag": False},
        {"ticker": "ZETA", "tier": 3, "defcon": 3, "flag": False},
    ]


def test_check_g_passes_on_canonical_sat_and_config(tmp_path):
    mod = _load_hook()
    cfg = _write_config_yaml(tmp_path, _build_canonical_sat_list_364())
    wb = _build_sat_wb_clean()
    err = mod._check_g_sparrate_sigma(wb, cfg)
    assert err is None, f"§G fail unexpected: {err}"


def test_check_g_fails_when_soll_mismatches_anker(tmp_path):
    mod = _load_hook()
    sats = _build_canonical_sat_list_364()
    sats.append({"ticker": "EXTRA", "tier": 1, "defcon": 3, "flag": False})  # SOLL +40 = 404 ≠ 364
    cfg = _write_config_yaml(tmp_path, sats)
    wb = _build_sat_wb_clean()
    err = mod._check_g_sparrate_sigma(wb, cfg)
    assert err is not None
    assert "SOLL" in err or "404" in err


def test_check_g_funded_echo_derives_from_config(tmp_path):
    """Funded/SOLL aus config abgeleitet, NICHT hardcoded — alternativer Roster (10× Tier 1 D3 → 400)."""
    mod = _load_hook()
    sats = [{"ticker": f"X{i}", "tier": 1, "defcon": 3, "flag": False} for i in range(10)]
    cfg = _write_config_yaml(tmp_path, sats, anker=400)
    wb = _build_sat_wb_clean(funded=400, soll=400)
    err = mod._check_g_sparrate_sigma(wb, cfg)
    assert err is None, f"§G fail bei alternativem Roster: {err}"


def test_check_g_fails_when_funded_echo_mismatches(tmp_path):
    """Funded-Echo zeigt 999 statt config-Funded 210 → §G FAIL (Display-Drift)."""
    mod = _load_hook()
    cfg = _write_config_yaml(tmp_path, _build_canonical_sat_list_364())
    wb = _build_sat_wb_clean(funded=999, soll=364)
    err = mod._check_g_sparrate_sigma(wb, cfg)
    assert err is not None
    assert "Funded" in err


def test_check_g_fails_on_missing_funded_echo(tmp_path):
    mod = _load_hook()
    cfg = _write_config_yaml(tmp_path, _build_canonical_sat_list_364())
    wb = _build_sat_wb_clean()
    wb["Satelliten Monitor"]["N20"] = None  # Funded-Echo entfernt
    err = mod._check_g_sparrate_sigma(wb, cfg)
    assert err is not None
    assert "Funded-Echo" in err


@pytest.mark.integration
def test_check_g_passes_on_live_config_yaml_and_canonical_wb():
    """Integrationstest — Live config.yaml (SOLL 364 / Funded 210) + canonical Sat-Workbook → PASS."""
    mod = _load_hook()
    cfg_path = _live_config_yaml_path()
    if not cfg_path.is_file():
        pytest.skip(f"Live config.yaml fehlt: {cfg_path}")
    wb = _build_sat_wb_clean()
    err = mod._check_g_sparrate_sigma(wb, cfg_path)
    assert err is None, f"§G fail gegen Live-config: {err}"


# ----- P13 Sheets-Tupel-Expansion -----


def test_p13_rebal_profile_includes_us_exposure_and_parameter():
    mod = _load_hook()
    sheets = mod._PROFILES["Rebalancing_Tool"]["sheets"]
    assert "Portfolio & Rebalancing" in sheets
    assert "US-Exposure" in sheets
    assert "Parameter & Regeln" in sheets


def test_p13_sat_profile_includes_quickscreen_ampel():
    mod = _load_hook()
    sheets = mod._PROFILES["Satelliten_Monitor"]["sheets"]
    assert "Satelliten Monitor" in sheets
    assert "QuickScreen Ampel" in sheets


def test_p13_watchlist_profile_unchanged():
    mod = _load_hook()
    sheets = mod._PROFILES["Watchlist_Ersatzbank_Monitor"]["sheets"]
    assert sheets == ()
