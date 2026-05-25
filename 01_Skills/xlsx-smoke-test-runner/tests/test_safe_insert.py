"""TDD-Tests für safe_insert.py (SPEC §4.1 Akzeptanz-Kriterien).

AMZN-Bug-Klasse (feedback_openpyxl_insert_merge_trap):
merged Range über insert-Zeile schluckt non-anchor-Writes silent.
"""

from __future__ import annotations

import openpyxl
import pytest
from safe_insert import safe_insert_cols, safe_insert_rows, safe_save


def _new_ws_with_merge(merge_range: str = "A2:C2"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "header-row1"
    ws["A2"] = "merged-anchor"
    ws["A3"] = "row3"
    ws["A4"] = "row4"
    ws.merge_cells(merge_range)
    return ws


def test_safe_insert_rows_preserves_overlapping_merge_shifted():
    ws = _new_ws_with_merge("A2:C2")
    safe_insert_rows(ws, idx=2, amount=1)

    merged = [str(r) for r in ws.merged_cells.ranges]
    assert "A3:C3" in merged, f"merge nicht um amount=1 verschoben: {merged}"
    assert ws["A3"].value == "merged-anchor"
    assert ws["A4"].value == "row3"


def test_safe_insert_rows_preserves_merge_with_amount_two():
    ws = _new_ws_with_merge("B2:D2")
    safe_insert_rows(ws, idx=2, amount=2)

    merged = [str(r) for r in ws.merged_cells.ranges]
    assert "B4:D4" in merged, f"merge nicht um amount=2 verschoben: {merged}"


def test_safe_insert_rows_expands_spanning_merge():
    """min_row < idx <= max_row → max_row += amount (Codex-R3 HIGH-3)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:C3")
    safe_insert_rows(ws, idx=2, amount=1)

    merged = [str(r) for r in ws.merged_cells.ranges]
    assert "A1:C4" in merged, f"spanning merge nicht expandiert: {merged}"


def test_safe_insert_cols_expands_spanning_merge():
    """col-spanning Analog (Codex-R3 HIGH-3)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:C3")
    safe_insert_cols(ws, idx=2, amount=1)

    merged = [str(r) for r in ws.merged_cells.ranges]
    assert "A1:D3" in merged, f"col-spanning merge nicht expandiert: {merged}"


def test_safe_insert_rows_does_not_touch_non_overlapping_merge():
    ws = _new_ws_with_merge("A2:C2")
    ws.merge_cells("A10:B10")
    safe_insert_rows(ws, idx=5, amount=1)

    merged = sorted(str(r) for r in ws.merged_cells.ranges)
    assert "A2:C2" in merged, f"vor-insert merge dürfen unverändert bleiben: {merged}"
    assert "A11:B11" in merged, f"nach-insert merge muss verschoben werden: {merged}"


def test_safe_insert_rows_amount_zero_raises_value_error():
    ws = _new_ws_with_merge()
    with pytest.raises(ValueError):
        safe_insert_rows(ws, idx=2, amount=0)


def test_safe_insert_rows_amount_negative_raises_value_error():
    ws = _new_ws_with_merge()
    with pytest.raises(ValueError):
        safe_insert_rows(ws, idx=2, amount=-1)


def test_safe_insert_rows_idx_zero_raises_index_error():
    ws = _new_ws_with_merge()
    with pytest.raises(IndexError):
        safe_insert_rows(ws, idx=0, amount=1)


def test_safe_insert_cols_preserves_overlapping_merge_shifted():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "col-a"
    ws["B1"] = "col-b"
    ws["C1"] = "col-c"
    ws.merge_cells("B1:B3")

    safe_insert_cols(ws, idx=2, amount=1)

    merged = [str(r) for r in ws.merged_cells.ranges]
    assert "C1:C3" in merged, f"col-merge nicht um amount=1 verschoben: {merged}"


def test_safe_insert_cols_idx_zero_raises_index_error():
    wb = openpyxl.Workbook()
    ws = wb.active
    with pytest.raises(IndexError):
        safe_insert_cols(ws, idx=0, amount=1)


def test_safe_insert_cols_amount_zero_raises_value_error():
    wb = openpyxl.Workbook()
    ws = wb.active
    with pytest.raises(ValueError):
        safe_insert_cols(ws, idx=2, amount=0)


def test_safe_save_clean_watchlist_returns_without_exception(tmp_path):
    wb = openpyxl.Workbook()
    wb.active["A1"] = "fixture"
    path = tmp_path / "Watchlist_Ersatzbank_Monitor_fresh.xlsx"
    safe_save(wb, path, profile="Watchlist_Ersatzbank_Monitor")
    assert path.is_file()


def test_safe_save_post_save_verify_fail_raises_runtime_error(tmp_path):
    wb = openpyxl.Workbook()
    wb.active["A1"] = None
    path = tmp_path / "Watchlist_Ersatzbank_Monitor_bad.xlsx"
    with pytest.raises(RuntimeError) as exc:
        safe_save(wb, path, profile="Watchlist_Ersatzbank_Monitor")
    assert "verify FAIL" in str(exc.value)
    assert path.is_file(), "wb.save lief, nur verify failed (kein Rollback in v0.1)"


def test_safe_insert_rows_raises_runtime_error_on_restore_fail(monkeypatch):
    """Codex-R3 HIGH-2: Re-Read-Assert (silent merge-restore-fail)."""
    ws = _new_ws_with_merge("A2:C2")
    original_merge = ws.merge_cells

    call_count = {"n": 0}

    def broken_merge(*args, **kwargs):
        call_count["n"] += 1
        # silent no-op (simuliert openpyxl-Edge-Case wo merge nicht persistiert)

    ws.merge_cells = broken_merge  # type: ignore[assignment]
    try:
        with pytest.raises(RuntimeError, match="merge restoration"):
            safe_insert_rows(ws, idx=2, amount=1)
    finally:
        ws.merge_cells = original_merge  # type: ignore[assignment]


def test_safe_insert_rows_does_not_unmerge_untouched_ranges():
    """Codex-R3 HIGH-1: nur betroffene Ranges werden unmerged (statt pauschal)."""
    ws = _new_ws_with_merge("A2:C2")
    ws.merge_cells("E1:F1")  # weit über idx, sollte unangetastet bleiben

    # Track unmerge calls
    unmerge_calls: list[str] = []
    original_unmerge = ws.unmerge_cells

    def tracking_unmerge(range_string):
        unmerge_calls.append(range_string)
        original_unmerge(range_string)

    ws.unmerge_cells = tracking_unmerge  # type: ignore[assignment]
    safe_insert_rows(ws, idx=5, amount=1)

    # E1:F1 ist max_row=1 < idx=5 → darf nicht unmerged werden
    assert "E1:F1" not in unmerge_calls, f"untouched merge wurde unnötig unmerged: {unmerge_calls}"
    merged = sorted(str(r) for r in ws.merged_cells.ranges)
    assert "E1:F1" in merged, f"untouched merge fehlt: {merged}"


def test_safe_insert_rows_round_trip_via_disk_keeps_merge(tmp_path):
    ws = _new_ws_with_merge("A2:C2")
    safe_insert_rows(ws, idx=2, amount=1)

    path = tmp_path / "round_trip.xlsx"
    ws.parent.save(path)
    ws.parent.close()

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    merged = [str(r) for r in ws2.merged_cells.ranges]
    assert "A3:C3" in merged, f"merge nach disk-round-trip fehlt: {merged}"
    assert ws2["A3"].value == "merged-anchor"
    wb2.close()
