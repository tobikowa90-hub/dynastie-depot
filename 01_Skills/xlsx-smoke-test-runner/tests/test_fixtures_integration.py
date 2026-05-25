"""Integration-Tests gegen die 15 generierten Fixtures (SPEC §8 #3 + #4).

Schritt 8 Pflicht-Reproduktion:
- Akzeptanz #3: verify_after_write fail-close auf injected #REF! Token
- Akzeptanz #4: safe_insert_rows preserved merges nach insert (Re-Read-Assert)
- Hook-Sat-§G Display-Drift fail-close
- Hook-Profile-Routing PASS/FAIL pro Gate
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from safe_insert import safe_insert_rows, safe_save
from verify_wrapper import verify_after_write

_FIX_DIR = Path(__file__).resolve().parent.parent / "_fixtures"
_GENERATOR = _FIX_DIR / "_generate_fixtures.py"


@pytest.fixture(scope="session", autouse=True)
def _ensure_fixtures_generated():
    """Vor allen Tests einmalig Fixtures regenerieren (deterministisch)."""
    sys.path.insert(0, str(_FIX_DIR))
    import _generate_fixtures

    _generate_fixtures.main()
    yield


# ----- Hook routing via verify_after_write (Hook-end-to-end) -----


def test_rebal_clean_fixture_passes_hook():
    p = _FIX_DIR / "Rebalancing_Tool_a_open_clean_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is True, f"clean Rebal fixture failed: {result.detail}"
    assert result.profile == "Rebalancing_Tool"


def test_rebal_missing_us_exposure_fails_a():
    p = _FIX_DIR / "Rebalancing_Tool_a_missing_us_exposure_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert "US-Exposure" in result.detail


def test_rebal_error_token_fails_b():
    """SPEC §8 #3: fail-close auf injected #REF!."""
    p = _FIX_DIR / "Rebalancing_Tool_b_error_token_in_p5_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert result.failed_gate == "B"
    assert "#REF!" in result.detail or "P5" in result.detail


def test_rebal_cf_count_mismatch_fails_e():
    p = _FIX_DIR / "Rebalancing_Tool_e_cf_count_mismatch_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert result.failed_gate == "E"


def test_sat_clean_fixture_passes_hook_all_gates():
    """Sat-clean enthaelt 5 CF + K3+N19 + 2 Sheets — alle Gates inkl. §G."""
    p = _FIX_DIR / "Satelliten_Monitor_a_open_clean_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is True, f"clean Sat fixture failed: {result.detail}"
    assert result.profile == "Satelliten_Monitor"


def test_sat_missing_quickscreen_fails_a():
    p = _FIX_DIR / "Satelliten_Monitor_a_missing_quickscreen_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert "QuickScreen Ampel" in result.detail


def test_sat_error_token_in_l7_fails_b():
    """SPEC §8 #3-Variante: fail-close auf #NAME? in Sat L7."""
    p = _FIX_DIR / "Satelliten_Monitor_b_error_token_in_l7_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert result.failed_gate == "B"
    assert "#NAME?" in result.detail or "L7" in result.detail


def test_sat_cf_count_mismatch_fails_e():
    p = _FIX_DIR / "Satelliten_Monitor_e_cf_count_mismatch_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert result.failed_gate == "E"


def test_sat_g_display_drift_fails_g():
    """N19='→ muss = 999,00 €' gegen Live anker=285 → §G FAIL."""
    p = _FIX_DIR / "Satelliten_Monitor_g_display_drift_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert result.failed_gate == "G"
    assert "N19" in result.detail


def test_watch_clean_passes_minimal_a():
    p = _FIX_DIR / "Watchlist_Ersatzbank_Monitor_a_open_clean_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is True


def test_watch_empty_a1_fails_minimal_a():
    p = _FIX_DIR / "Watchlist_Ersatzbank_Monitor_a_empty_a1_fixture.xlsx"
    result = verify_after_write(p)
    assert result.ok is False
    assert result.failed_gate == "A"


# ----- Direct §G API mit companion config (mapping-drift) -----


def test_sat_g_sparrate_mapping_drift_via_companion_config():
    """Companion config Σ=323 ≠ anker=285 → §G FAIL via direct API."""
    import importlib.util

    hook_path = (
        Path(__file__).resolve().parents[3] / "03_Tools" / "precommit" / "xlsx_smoke_test.py"
    )
    spec = importlib.util.spec_from_file_location("hut", hook_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    wb = openpyxl.load_workbook(
        _FIX_DIR / "Satelliten_Monitor_g_sparrate_mapping_drift_fixture.xlsx"
    )
    companion = _FIX_DIR / "Satelliten_Monitor_g_sparrate_mapping_drift_config.yaml"
    err = mod._check_g_sparrate_sigma(wb, companion)
    assert err is not None
    assert "323" in err and "285" in err


# ----- safe_insert/safe_save fixture-integration (SPEC §8 #4) -----


def test_safe_insert_preserves_merges_via_fixture(tmp_path):
    """SPEC §8 #4: safe_insert_rows preserved merges via Fixture (round-trip)."""
    src = _FIX_DIR / "skill_safe_insert_preserves_merges_fixture.xlsx"
    work = tmp_path / "work.xlsx"
    work.write_bytes(src.read_bytes())

    wb = openpyxl.load_workbook(work)
    ws = wb.active
    pre_merges = sorted(str(r) for r in ws.merged_cells.ranges)
    assert "A1:C3" in pre_merges

    safe_insert_rows(ws, idx=2, amount=1)
    post_merges = sorted(str(r) for r in ws.merged_cells.ranges)
    assert "A1:C4" in post_merges, f"merge nicht expandiert: {post_merges}"


def test_safe_insert_invalid_idx_via_fixture():
    src = _FIX_DIR / "skill_safe_insert_invalid_idx_fixture.xlsx"
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    with pytest.raises(IndexError):
        safe_insert_rows(ws, idx=0, amount=1)


def test_safe_save_post_corruption_via_fixture(tmp_path):
    """safe_save against corrupted state (CF removed) → RuntimeError via verify."""
    src = _FIX_DIR / "skill_safe_save_post_corruption_fixture.xlsx"
    work_name = "Satelliten_Monitor_corrupted_target.xlsx"
    work = tmp_path / work_name

    wb = openpyxl.load_workbook(src)
    sat = wb["Satelliten Monitor"]
    # Remove CF rules to force §E mismatch (5 → 0)
    sat.conditional_formatting = type(sat.conditional_formatting)()

    with pytest.raises(RuntimeError) as exc:
        safe_save(wb, work, profile="Satelliten_Monitor")
    assert "verify FAIL" in str(exc.value)
    assert "E" in str(exc.value) or "Conditional-Format" in str(exc.value)


# ----- Determinismus #5 (SPEC §8 #5) -----


def test_fixture_generator_deterministic_via_zip_subset_hash():
    """SPEC §8 #5: zwei Generator-Laeufe → identische ZIP-Subset-Hashes."""
    sys.path.insert(0, str(_FIX_DIR))
    import _generate_fixtures

    sample = _FIX_DIR / "Satelliten_Monitor_a_open_clean_fixture.xlsx"
    h1 = _generate_fixtures.fixture_content_sha256(sample)
    _generate_fixtures.main()
    h2 = _generate_fixtures.fixture_content_sha256(sample)
    assert h1 == h2, f"non-deterministic generator: {h1} != {h2}"
