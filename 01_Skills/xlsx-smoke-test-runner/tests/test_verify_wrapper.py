"""TDD-Tests für verify_wrapper.py (SPEC §4.2 Akzeptanz-Kriterien)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from verify_wrapper import (
    VerifyResult,
    _load_hook_module,
    verify_after_write,
    verify_batch,
)


def _make_clean_watchlist(tmp_path: Path) -> Path:
    """Watchlist-Minimal-Profil: A1 gesetzt → Hook PASS."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "fixture"
    p = tmp_path / "Watchlist_Ersatzbank_Monitor_clean.xlsx"
    wb.save(p)
    wb.close()
    return p


def _make_empty_a1_watchlist(tmp_path: Path) -> Path:
    """Watchlist-Minimal-Profil: A1=None → Hook FAIL Punkt A."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = None
    p = tmp_path / "Watchlist_Ersatzbank_Monitor_empty_a1.xlsx"
    wb.save(p)
    wb.close()
    return p


def _make_unknown_profile(tmp_path: Path) -> Path:
    """Filename matches kein Profil-Prefix → Hook FAIL UNKNOWN."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    p = tmp_path / "Unbekannter_Name.xlsx"
    wb.save(p)
    wb.close()
    return p


def test_load_hook_module_returns_module_with_validate_file():
    mod = _load_hook_module()
    assert hasattr(mod, "validate_file"), "Hook-Modul ohne validate_file"
    assert callable(mod.validate_file)


def test_verify_result_is_frozen_dataclass():
    import dataclasses

    r = VerifyResult(ok=True, profile="X", failed_gate=None, detail="")
    with pytest.raises(dataclasses.FrozenInstanceError):
        r.ok = False  # type: ignore[misc]


def test_verify_after_write_clean_watchlist_returns_ok(tmp_path):
    path = _make_clean_watchlist(tmp_path)
    result = verify_after_write(path)
    assert isinstance(result, VerifyResult)
    assert result.ok is True
    assert result.profile == "Watchlist_Ersatzbank_Monitor"
    assert result.failed_gate is None


def test_verify_after_write_empty_a1_returns_not_ok(tmp_path):
    path = _make_empty_a1_watchlist(tmp_path)
    result = verify_after_write(path)
    assert result.ok is False
    assert result.profile == "Watchlist_Ersatzbank_Monitor"
    # Hook fail-message enthält 'Punkt A' für minimal-A1
    assert "Punkt A" in result.detail or "A1" in result.detail
    # Codex-R3 MED-2: failed_gate-Parse expliziert
    assert result.failed_gate == "A", f"failed_gate={result.failed_gate!r}, expected 'A'"


def test_verify_after_write_run_g_flag_is_noop_v01(tmp_path):
    """Codex-R3 MED-3: run_g True vs False ergibt identisches Ergebnis in v0.1."""
    path = _make_clean_watchlist(tmp_path)
    r_true = verify_after_write(path, run_g=True)
    r_false = verify_after_write(path, run_g=False)
    assert r_true.ok == r_false.ok
    assert r_true.profile == r_false.profile
    assert r_true.failed_gate == r_false.failed_gate


def test_verify_after_write_unknown_profile_returns_not_ok(tmp_path):
    path = _make_unknown_profile(tmp_path)
    result = verify_after_write(path)
    assert result.ok is False
    assert result.failed_gate is not None  # mind. ein Gate-Marker gesetzt


def test_verify_after_write_missing_file_raises_io_error(tmp_path):
    missing = tmp_path / "does_not_exist.xlsx"
    with pytest.raises(FileNotFoundError):
        verify_after_write(missing)


def test_verify_batch_processes_all_even_on_intermediate_fail(tmp_path):
    ok_path = _make_clean_watchlist(tmp_path)
    bad_path = _make_empty_a1_watchlist(tmp_path)
    results = verify_batch([ok_path, bad_path, ok_path])
    assert len(results) == 3
    assert results[0].ok is True
    assert results[1].ok is False
    assert results[2].ok is True
