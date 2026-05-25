"""safe_insert.py — openpyxl Insert+Merge-Trap-Schutz (SPEC §4.1).

AMZN-Bug-Klasse: merged Range über insert-Zeile schluckt non-anchor-Writes
silent (Memory feedback_openpyxl_insert_merge_trap).

Algorithmus (rows-Variante, Codex-R3-HIGH-1+HIGH-2 final):
1. Klassifikation pro merged-range relativ zu (idx, amount):
   - max_row < idx              → unangetastet (NICHT unmerge)
   - min_row >= idx             → shift   (new = +amount)
   - min_row < idx <= max_row   → expand  (max_row += amount)
2. Capture + unmerge NUR betroffene ranges (selective, Codex-HIGH-1)
3. ws.insert_rows(idx, amount)
4. Re-merge mit neuen bounds
5. POST-CONDITION (Codex-HIGH-2): Re-Read-Assert — erwartete Ranges müssen in
   ws.merged_cells.ranges sein, sonst RuntimeError.

cols-Variante symmetrisch.

Read-only-Vertrag: schreibt NICHT die Datei (Caller muss wb.save()).
"""

from __future__ import annotations

from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _validate_args(idx: int, amount: int) -> None:
    if amount < 1:
        raise ValueError(f"amount must be >= 1, got {amount}")
    if idx < 1:
        raise IndexError(f"idx must be >= 1 (openpyxl is 1-indexed), got {idx}")


def _classify_rows(min_row: int, max_row: int, idx: int, amount: int) -> tuple[int, int] | None:
    """Returns (new_min_row, new_max_row) für betroffene Ranges, None für untouched."""
    if max_row < idx:
        return None
    if min_row >= idx:
        return (min_row + amount, max_row + amount)
    return (min_row, max_row + amount)


def _classify_cols(min_col: int, max_col: int, idx: int, amount: int) -> tuple[int, int] | None:
    """Returns (new_min_col, new_max_col) für betroffene Ranges, None für untouched."""
    if max_col < idx:
        return None
    if min_col >= idx:
        return (min_col + amount, max_col + amount)
    return (min_col, max_col + amount)


def _assert_merges_restored(ws: Worksheet, expected: list[str]) -> None:
    """Re-Read-Assert (Codex-HIGH-2): RuntimeError bei fehlenden Re-Merges."""
    actual = {str(r) for r in ws.merged_cells.ranges}
    missing = [e for e in expected if e not in actual]
    if missing:
        raise RuntimeError(
            f"safe_insert: merge restoration silent fail — "
            f"missing={missing}, actual={sorted(actual)}"
        )


def safe_insert_rows(ws: Worksheet, idx: int, amount: int = 1) -> None:
    """Insert `amount` Zeilen ab Zeile `idx` mit Merge-Preservation.

    Wirft:
        ValueError bei amount < 1
        IndexError bei idx < 1
        RuntimeError falls Merge-Restoration silent fail (Re-Read-Assert)
    """
    _validate_args(idx, amount)

    affected: list[tuple[int, int, int, int]] = []  # (new_min_row, new_max_row, min_col, max_col)
    for r in list(ws.merged_cells.ranges):
        new_rows = _classify_rows(r.min_row, r.max_row, idx, amount)
        if new_rows is None:
            continue
        affected.append((new_rows[0], new_rows[1], r.min_col, r.max_col))
        ws.unmerge_cells(str(r))

    ws.insert_rows(idx, amount)

    expected: list[str] = []
    for new_min_row, new_max_row, min_col, max_col in affected:
        ws.merge_cells(
            start_row=new_min_row,
            end_row=new_max_row,
            start_column=min_col,
            end_column=max_col,
        )
        expected.append(
            f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
        )

    _assert_merges_restored(ws, expected)


def safe_insert_cols(ws: Worksheet, idx: int, amount: int = 1) -> None:
    """Insert `amount` Spalten ab Spalte `idx` mit Merge-Preservation.

    Wirft:
        ValueError bei amount < 1
        IndexError bei idx < 1
        RuntimeError falls Merge-Restoration silent fail
    """
    _validate_args(idx, amount)

    affected: list[tuple[int, int, int, int]] = []  # (min_row, max_row, new_min_col, new_max_col)
    for r in list(ws.merged_cells.ranges):
        new_cols = _classify_cols(r.min_col, r.max_col, idx, amount)
        if new_cols is None:
            continue
        affected.append((r.min_row, r.max_row, new_cols[0], new_cols[1]))
        ws.unmerge_cells(str(r))

    ws.insert_cols(idx, amount)

    expected: list[str] = []
    for min_row, max_row, new_min_col, new_max_col in affected:
        ws.merge_cells(
            start_row=min_row,
            end_row=max_row,
            start_column=new_min_col,
            end_column=new_max_col,
        )
        expected.append(
            f"{get_column_letter(new_min_col)}{min_row}:{get_column_letter(new_max_col)}{max_row}"
        )

    _assert_merges_restored(ws, expected)


def safe_save(
    wb: openpyxl.Workbook,
    path: Path,
    profile: Literal[
        "Rebalancing_Tool",
        "Satelliten_Monitor",
        "Watchlist_Ersatzbank_Monitor",
    ],
) -> None:
    """Save + Post-Save-Verify via Hook-Funktionen (SPEC §4.1).

    Delegiert an verify_wrapper.verify_after_write(path); wirft
    RuntimeError mit Hook-Fehlerdetail bei rc != 0.

    Idempotent — kein Rollback bei verify-fail (Caller-Verantwortung).
    """
    from verify_wrapper import verify_after_write  # lokal um circular avoid

    wb.save(path)
    result = verify_after_write(Path(path))
    if not result.ok:
        raise RuntimeError(
            f"safe_save verify FAIL [{path}|profile={profile}|"
            f"failed_gate={result.failed_gate}]: {result.detail}"
        )


__all__ = [
    "safe_insert_cols",
    "safe_insert_rows",
    "safe_save",
]
