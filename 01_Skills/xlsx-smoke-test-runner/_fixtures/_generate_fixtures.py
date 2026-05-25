"""Deterministischer Fixture-Generator für xlsx-smoke-test-runner (SPEC §5.1+§5.2).

15 Fixtures = 12 Profile-Fixtures (Rebal 4 + Sat 6 + Watch 2) + 3 Skill-Helper (safe_insert/safe_save).

- Beschreibende Namen (NICHT bad_*/good_* per Memory feedback_cr_convergence_and_project_compat).
- Profile-Fixture-Namen starten mit Profil-Prefix (Hook _resolve_profil match).
- Skill-Helper-Fixtures ohne Prefix (für direkte API-Tests).
- Pflicht-Mitigation Codex-R2-MED-1: ZIP-Subset-Hashing exkludiert
  docProps/{core,app}.xml (volatile Timestamps/Creator).
- Idempotent: kein Timestamp/Random.

Lauf: `python 01_Skills/xlsx-smoke-test-runner/_fixtures/_generate_fixtures.py`
"""

from __future__ import annotations

import hashlib
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

_HERE = Path(__file__).resolve().parent
_EXCLUDED_ZIP_MEMBERS = frozenset({"docProps/core.xml", "docProps/app.xml"})


def fixture_content_sha256(path: Path) -> str:
    """SHA256 ueber ZIP-Subset (excl. docProps-Volatile) — deterministisch (SPEC §5.2)."""
    h = hashlib.sha256()
    with zipfile.ZipFile(path, "r") as zf:
        for name in sorted(zf.namelist()):
            if name in _EXCLUDED_ZIP_MEMBERS:
                continue
            h.update(name.encode("utf-8"))
            h.update(zf.read(name))
    return h.hexdigest()


def _save(wb: openpyxl.Workbook, name: str) -> Path:
    path = _HERE / name
    wb.save(path)
    print(
        f"  wrote {name} ({path.stat().st_size} bytes, sha256={fixture_content_sha256(path)[:12]})"
    )
    return path


def _add_cf_rules(ws, count: int, start_row: int = 5) -> None:
    """Adds `count` Cell-Is-Rules to ws for §E-CF-count-Tests (deterministisch)."""
    red_fill = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
    for i in range(count):
        rule = CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=red_fill)
        ws.conditional_formatting.add(f"A{start_row + i}:A{start_row + i}", rule)


# ----- 7a Basis-Fixtures (12 Profile-Fixtures) -----


def _build_rebal_a_open_clean() -> None:
    """3-Sheet-Rebal + 6 CFs + no error tokens (Hook PASS §A+§B+§E)."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Portfolio & Rebalancing"
    ws1["A1"] = "Header"
    _add_cf_rules(ws1, 5)
    ws2 = wb.create_sheet("US-Exposure")
    _add_cf_rules(ws2, 0)
    ws3 = wb.create_sheet("Parameter & Regeln")
    _add_cf_rules(ws3, 1)
    _save(wb, "Rebalancing_Tool_a_open_clean_fixture.xlsx")


def _build_rebal_a_missing_us_exposure() -> None:
    """Nur 2 Sheets (US-Exposure fehlt) → Hook FAIL §A."""
    wb = openpyxl.Workbook()
    wb.active.title = "Portfolio & Rebalancing"
    wb.create_sheet("Parameter & Regeln")
    _save(wb, "Rebalancing_Tool_a_missing_us_exposure_fixture.xlsx")


def _build_rebal_b_error_token_in_p5() -> None:
    """P5 enthaelt '#REF!' → Hook FAIL §B."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio & Rebalancing"
    ws["P5"] = "#REF!"
    wb.create_sheet("US-Exposure")
    wb.create_sheet("Parameter & Regeln")
    _add_cf_rules(ws, 6)
    _save(wb, "Rebalancing_Tool_b_error_token_in_p5_fixture.xlsx")


def _build_rebal_e_cf_count_mismatch() -> None:
    """5 CFs statt 6 → Hook FAIL §E."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio & Rebalancing"
    wb.create_sheet("US-Exposure")
    wb.create_sheet("Parameter & Regeln")
    _add_cf_rules(ws, 5)  # nur 5, erwartet 6
    _save(wb, "Rebalancing_Tool_e_cf_count_mismatch_fixture.xlsx")


def _build_sat_canonical_wb(cf_count: int = 5) -> openpyxl.Workbook:
    """Sat-Workbook mit erwarteten Sheets + K3/N19-Display (PASS §G default)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Satelliten Monitor"
    ws["K3"] = "Ergebnis: ● 7 Voll  ● 1 D2-Sockelbetrag (V)  ● 4 Eingefroren"
    ws["N19"] = "→ muss = 285,00 €"
    _add_cf_rules(ws, cf_count)
    wb.create_sheet("QuickScreen Ampel")
    return wb


def _build_sat_a_open_clean() -> None:
    """2-Sheet-Sat + 5 CFs + valid §G display (Hook PASS §A+§B+§E+§G)."""
    wb = _build_sat_canonical_wb()
    _save(wb, "Satelliten_Monitor_a_open_clean_fixture.xlsx")


def _build_sat_a_missing_quickscreen() -> None:
    """QuickScreen Ampel fehlt → Hook FAIL §A."""
    wb = openpyxl.Workbook()
    wb.active.title = "Satelliten Monitor"
    _save(wb, "Satelliten_Monitor_a_missing_quickscreen_fixture.xlsx")


def _build_sat_b_error_token_in_l7() -> None:
    """L7 enthaelt '#NAME?' → Hook FAIL §B."""
    wb = _build_sat_canonical_wb()
    wb["Satelliten Monitor"]["L7"] = "#NAME?"
    _save(wb, "Satelliten_Monitor_b_error_token_in_l7_fixture.xlsx")


def _build_sat_e_cf_count_mismatch() -> None:
    """4 CFs statt 5 → Hook FAIL §E."""
    wb = _build_sat_canonical_wb(cf_count=4)
    _save(wb, "Satelliten_Monitor_e_cf_count_mismatch_fixture.xlsx")


def _build_sat_g_sparrate_mapping_drift() -> None:
    """xlsx clean (PASS §A+§B+§E+§G-display) + companion config mit Σ=323 ≠ 285."""
    wb = _build_sat_canonical_wb()
    _save(wb, "Satelliten_Monitor_g_sparrate_mapping_drift_fixture.xlsx")

    # Companion config: APH flag=False statt True → Σ +19 = 304
    # plus AVGO flag=False → Σ +19 = 323
    drift_sats = [
        {"ticker": "ASML", "defcon": 3, "flag": False},
        {"ticker": "AVGO", "defcon": 2, "flag": False},  # drift: flag was True
        {"ticker": "MSFT", "defcon": 2, "flag": True},
        {"ticker": "COST", "defcon": 3, "flag": False},
        {"ticker": "RMS", "defcon": 3, "flag": False},
        {"ticker": "VEEV", "defcon": 3, "flag": False},
        {"ticker": "SU", "defcon": 3, "flag": False},
        {"ticker": "BRK.B", "defcon": 3, "flag": False},
        {"ticker": "V", "defcon": 2, "flag": False},
        {"ticker": "TMO", "defcon": 3, "flag": False},
        {"ticker": "APH", "defcon": 2, "flag": False},  # drift: flag was True
        {"ticker": "AMZN", "defcon": 1, "flag": True},
    ]
    cfg = {"brokers": {"scalable": {"sparrate_eur": 285}}, "satelliten": drift_sats}
    cfg_path = _HERE / "Satelliten_Monitor_g_sparrate_mapping_drift_config.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    print(f"  wrote {cfg_path.name} (companion config, sigma=323)")


def _build_sat_g_display_drift() -> None:
    """N19 = '→ muss = 999,00 €' (display drift) → §G FAIL gegen Live anker."""
    wb = _build_sat_canonical_wb()
    wb["Satelliten Monitor"]["N19"] = "→ muss = 999,00 €"
    _save(wb, "Satelliten_Monitor_g_display_drift_fixture.xlsx")


def _build_watch_a_open_clean() -> None:
    """1-Sheet Watchlist + A1 gesetzt → Hook PASS §A minimal."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Watchlist"
    _save(wb, "Watchlist_Ersatzbank_Monitor_a_open_clean_fixture.xlsx")


def _build_watch_a_empty_a1() -> None:
    """A1 leer → Hook FAIL §A minimal."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = None
    _save(wb, "Watchlist_Ersatzbank_Monitor_a_empty_a1_fixture.xlsx")


# ----- 7c Skill-Helper Fixtures (3) -----


def _build_g_safe_insert_preserves_merges() -> None:
    """merged Range A1:C3, expected: nach safe_insert_rows(idx=2) → A1:C4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:C3")
    ws["A1"] = "merged-anchor"
    _save(wb, "skill_safe_insert_preserves_merges_fixture.xlsx")


def _build_g_safe_insert_invalid_idx() -> None:
    """Workbook für IndexError-Demo (safe_insert_rows(ws, idx=0) wirft IndexError)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "for invalid_idx test"
    _save(wb, "skill_safe_insert_invalid_idx_fixture.xlsx")


def _build_g_safe_save_post_corruption() -> None:
    """Sat-konformer Workbook für safe_save-Test mit injizierter post-save CF-corruption-Sim.

    Tatsaechliche Korruption findet im Test statt (nach wb.save in tmp_path);
    Fixture liefert nur den Baseline-State.
    """
    wb = _build_sat_canonical_wb()
    _save(wb, "skill_safe_save_post_corruption_fixture.xlsx")


def main() -> int:
    print("generating fixtures…")
    # 7a Basis-Fixtures (Rebal 4 + Sat 6 + Watch 2 = 12)
    _build_rebal_a_open_clean()
    _build_rebal_a_missing_us_exposure()
    _build_rebal_b_error_token_in_p5()
    _build_rebal_e_cf_count_mismatch()
    _build_sat_a_open_clean()
    _build_sat_a_missing_quickscreen()
    _build_sat_b_error_token_in_l7()
    _build_sat_e_cf_count_mismatch()
    _build_sat_g_sparrate_mapping_drift()  # 7b §G (1/2 + config.yaml)
    _build_sat_g_display_drift()  # 7b §G (2/2)
    _build_watch_a_open_clean()
    _build_watch_a_empty_a1()
    # 7c Skill-Helper Fixtures (3)
    _build_g_safe_insert_preserves_merges()
    _build_g_safe_insert_invalid_idx()
    _build_g_safe_save_post_corruption()
    print("OK - 15 xlsx + 1 companion config.yaml generated")
    return 0


if __name__ == "__main__":
    sys.exit(main())
