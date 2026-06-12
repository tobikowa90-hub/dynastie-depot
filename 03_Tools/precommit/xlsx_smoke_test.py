"""xlsx-smoke-test pre-commit-Hook (Spec §3.1).

Reproduziert das §18.7-Verhalten (`03_Tools/xlsx-smoke-test.md`) fail-close als
git-Gate. Dieser Hook ist **Executor, nicht SSoT** — die Soll-Werte
(Sheet-Namen, Conditional-Format-Counts, Voll-vs-Minimal-Scope) leben in
`03_Tools/xlsx-smoke-test.md`. Implementiert den Excel-Desktop-Fallback
Punkt A (Open) + B (Error-Scan) + E (CF-Count); F/C/D = manuell/UI-only,
nicht git-gate-fähig.

Read-only: lädt nur via openpyxl, schreibt nie (Memory
feedback_onedrive_edit_collision — Edit-Collision bei offenem Editor).

Aufruf: pre-commit übergibt die staged Pfade als argv (`language: system`).
Fail-close: EXIT 1 bei erstem Fail, Pfad+Profil+Grund auf stderr.
"""

from __future__ import annotations

import contextlib
import re
import sys
from pathlib import Path

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

import openpyxl
import yaml
from openpyxl.utils.exceptions import InvalidFileException

# Error-Strings (xlsx-smoke-test.md Punkt B). Substring-Scan statt exact-match
# der md-Fallback-B: fängt sowohl reine '#REF!'-Werte als auch in Formeln
# eingebettete '=A1+#REF!' (data_only=False) → robustere Intent-Erfüllung.
# Bewusste Abweichung von der md-exact-match, zur Codex-Disposition (Spec §6).
_ERROR_TOKENS: tuple[str, ...] = ("#REF!", "#NAME?", "#VALUE!", "#N/A")

# Config-Anker für Punkt G (SPEC §4.3 Variante G, drift-doc §2.4).
# Anker-Pfad: brokers.scalable.sparrate_eur (verified config.yaml L27,
# 2026-05-25 post Codex-R1 HIGH-1 — portfolio.satelliten_sparrate existiert NICHT).
_CONFIG_YAML_PATH = (
    Path(__file__).resolve().parents[2] / "01_Skills" / "dynastie-depot" / "config.yaml"
)


def _echo_int_numbers(text: str) -> set[int]:
    """Extrahiert exakte ganzzahlige Zahlenwerte aus einem Funded/SOLL-Echo-String.

    Robuster Ersatz für den naiven Substring-Scan (Codex-Review 2026-06-08 HIGH):
    `funded_i in val` lieferte False-PASS, weil Substrings matchen ('210' ⊂ '2100',
    '364' ⊂ '3640'). Ein simpler Label-Anker ('Zahl direkt nach Funded') scheitert
    am realen B27-Format 'Funded Σ = 40+80+90 = 210€' (läse 40 statt 210).
    Lösung: ALLE Zahlen-Tokens word-boundary-exakt extrahieren (Lookarounds
    verhindern Teil-Matches in '2100'), deutsche Dezimal-/Tausenderform
    normalisieren ('210,00'→210, '1.500'→1500), auf int reduzieren. Der §G-Check
    prüft dann exakte Set-Membership (funded ∈ Zahlen, soll ∈ Zahlen) statt
    Substring — fängt jede Ziffern-Drift, ohne an den 4 realen Echo-Formaten
    (B3/K3/N20/B27) false-zu-failen. Korrektheit > Laufzeit.
    """
    out: set[int] = set()
    for m in re.finditer(r"(?<![\d.,])\d[\d.]*(?:,\d+)?(?![\d.,])", text):
        intpart = m.group(0).split(",")[0].replace(".", "")
        if intpart.isdigit():
            out.add(int(intpart))
    return out


# Profil-Soll-Werte. SSoT = 03_Tools/xlsx-smoke-test.md (Scope-Tabelle).
# `cf_rule_count` ist gegen die Live-xlsx beim ersten pre-commit-Smoke zu
# verifizieren/justieren (Spec Akzeptanz-Kriterium 1) — Annahme explizit (§0.1).
# `sheets` Tupel erweitert per P13 (drift-doc §1.1 + §2.1 — 2026-05-25 Stage-2 Schritt 6).
_PROFILES: dict[str, dict] = {
    "Rebalancing_Tool": {
        "scope": "voll",
        "sheets": (
            "Portfolio & Rebalancing",
            "US-Exposure",  # P13 NEU 2026-05-25 (drift-doc §1.1)
            "Parameter & Regeln",  # P13 NEU 2026-05-25 (drift-doc §1.1)
        ),
        "cf_rule_count": 9,  # CF-Ranges (sqref-Bloecke). exUSA-Re-Add 2026-06-12 7->9: User-Edit fuegte 2 CFs hinzu (K5:K24 dataBar 'Abw.%' + H25 expression 'GESAMT Ziel-Anteil-Check'), bestaetigt gewollt; +EXUS-Row-Insert verschob 6 bestehende Ranges (N11:N23->N12:N24 etc.). Vorher 7 (Live-State 2026-06-06 v4.0-Roster); davor 6 (2026-05-25). HINWEIS: K-dataBar nutzt x14-Extension -> openpyxl-Write (Score-Sync) wuerde sie strippen.
    },
    "Satelliten_Monitor": {
        "scope": "voll",
        "sheets": (
            "Satelliten Monitor",
            "QuickScreen Ampel",  # P13 NEU 2026-05-25 (drift-doc §2.1)
        ),
        "cf_rule_count": 11,  # 5 (Satelliten Monitor: L/M/N/O/R) + 6 (QuickScreen Ampel: D/D-G/E/F/G/H Ampel-Coloring, Umstrukturierung 2026-06-07) = 11 Ranges. Vorher 5 (QuickScreen hatte 0 CF).
    },
    "Watchlist_Ersatzbank_Monitor": {
        "scope": "minimal",  # md Minimal-Check-Annex: nur A1 + Existenz
        "sheets": (),
        "cf_rule_count": 0,
    },
}


def _fail(path: Path, profil: str, grund: str) -> int:
    print(
        f"❌ xlsx-smoke-test: FAIL [{path.name} | profil={profil}] {grund}",
        file=sys.stderr,
    )
    return 1


def _resolve_profil(path: Path) -> tuple[str, dict] | None:
    # NOTE (Codex-R3 MED-1 deferred): startswith ist kollision-anfällig bei
    # Backup-Suffixen (z.B. `Rebalancing_Tool_v3.4_backup.xlsx`). v0.2-Refactor-Pfad
    # → regex full-match `^<Profil>_v\d+\.\d+\.xlsx$` (SPEC §4.4 Codex-R1 LOW).
    # v0.1: bewusst startswith — kein Backup-Pattern aktuell im Repo.
    for key, prof in _PROFILES.items():
        if path.name.startswith(key):
            return key, prof
    return None


def _count_cf_rules(wb: openpyxl.Workbook) -> int:
    total = 0
    for ws in wb.worksheets:
        total += sum(1 for _ in ws.conditional_formatting)
    return total


def _derive_rate_eur(satellit_cfg: dict, tier_raten: dict) -> float:
    """Punkt G Mapping: config.yaml satellit → effektive (modulierte) Tier-Rate-EUR.

    Tier-Modell (Umstrukturierung 2026-06-07, löst flaches 38/19-Modell ab):
        effektive Rate = tier_basis × DEFCON-Modulation × FLAG
    - tier_basis = tier_raten[satellit.tier]   (Tier 1=40, 2=32, 3=18)
    - flag=True                     → 0           (heilig, überschreibt DEFCON score-unabhängig)
    - flag=False + defcon in {3,4}  → tier_basis        (volle Tier-Rate)
    - flag=False + defcon == 2      → tier_basis / 2    (halber Sockelbetrag)
    - flag=False + defcon == 1      → 0

    Wirft ValueError bei unmapped tier oder defcon.
    """
    tier = satellit_cfg.get("tier")
    if tier not in tier_raten:
        raise ValueError(f"unmapped tier={tier!r} für satellit={satellit_cfg.get('ticker', '?')}")
    base = tier_raten[tier]
    if satellit_cfg.get("flag") is True:
        return 0
    defcon = satellit_cfg.get("defcon")
    if defcon in (3, 4):
        return base
    if defcon == 2:
        return base / 2
    if defcon == 1:
        return 0
    raise ValueError(f"unmapped defcon={defcon!r} für satellit={satellit_cfg.get('ticker', '?')}")


def _check_g_sparrate_sigma(wb: openpyxl.Workbook, config_yaml_path: Path) -> str | None:
    """Punkt G Sparrate-Σ-Sanity (SPEC §4.3, Tier-Modell Umstrukturierung 2026-06-07).

    Steps:
    1. Load config_yaml_path (yaml.safe_load)
    2. tier_raten = cfg["satelliten_tier_raten"]
    3. SOLL-Σ = sum(tier_raten[s["tier"]] for s in satelliten)   [strukturell, stabil — bricht NICHT bei FLAG-Events]
    4. Assert SOLL-Σ == cfg["brokers"]["scalable"]["sparrate_eur"]  (Anker 364€)
    5. Funded-Σ = sum(_derive_rate_eur(s, tier_raten) ...)   [transient, moduliert]
    6. Assert "Satelliten Monitor" K3-Text enthält ' Voll' + 'Sockelbetrag' + 'Eingefroren'
    7. Assert Funded-Echo-Zelle (Content-Scan 'Funded'+'SOLL', layout-robust) enthält
       Funded-Σ + SOLL-Σ formatiert (deutsche Schreibweise mit Komma).

    Returns None bei PASS, str (Fehler-Detail) bei FAIL.
    """
    try:
        with Path(config_yaml_path).open(encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
    except (OSError, yaml.YAMLError) as e:
        return f"config.yaml unreadable: {e}"

    try:
        tier_raten = cfg["satelliten_tier_raten"]
        soll = sum(tier_raten[s["tier"]] for s in cfg["satelliten"])
        funded = sum(_derive_rate_eur(s, tier_raten) for s in cfg["satelliten"])
    except (KeyError, TypeError, ValueError) as e:
        return f"satelliten/tier-mapping fail: {e}"

    try:
        anker = cfg["brokers"]["scalable"]["sparrate_eur"]
    except (KeyError, TypeError) as e:
        return f"anker brokers.scalable.sparrate_eur missing: {e}"

    if soll != anker:
        return f"SOLL-Σ-Drift: derived={soll}€ != anker={anker}€ (config.yaml.brokers.scalable.sparrate_eur)"

    if "Satelliten Monitor" not in wb.sheetnames:
        return "Sheet 'Satelliten Monitor' fehlt — kein §G-Display-Check möglich"
    sat_ws = wb["Satelliten Monitor"]

    k3 = sat_ws["K3"].value or ""
    for needle in (" Voll", "Sockelbetrag", "Eingefroren"):
        if needle not in str(k3):
            return f"K3-Display-Drift: '{needle}' fehlt in {k3!r}"

    # Funded-Echo layout-robust per Content-Scan (alle Zellen mit 'Funded'+'SOLL'), statt fixe Zelle.
    # Begründung: fixe Zelle (vormals N19) driftet bei Roster-Resize (Umstrukturierung
    # 2026-06-07: N19 wurde Ticker-Zeile, Echo wanderte nach N20). Es gibt mehrere Echo-Zellen
    # (Header-Zeile + Detail-Footer) in unterschiedlichem Format ("210 €" vs "210,00 €") —
    # JEDE muss mit dem config-abgeleiteten Funded-/SOLL-Wert übereinstimmen (Drift in einer
    # Echo-Zelle = Fail). Match via exakte Zahlen-Token-Set-Membership (_echo_int_numbers),
    # NICHT Substring — verhindert False-PASS ('210' ⊂ '2100'). Werte aus config abgeleitet,
    # NICHT hardcoded. (Codex-Review 2026-06-08 HIGH.)
    echoes = [
        (c.coordinate, c.value)
        for row in sat_ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and "Funded" in c.value and "SOLL" in c.value
    ]
    if not echoes:
        return "Funded-Echo-Zelle (enthält 'Funded'+'SOLL') fehlt im Satelliten-Monitor (Punkt G)"
    funded_i = int(funded)
    soll_i = int(anker)
    for coord, val in echoes:
        nums = _echo_int_numbers(val)
        if funded_i not in nums:
            return f"Funded-Echo-Drift {coord}: erwartet Funded {funded_i} € (exakte Zahl), live={val!r}"
        if soll_i not in nums:
            return (
                f"Funded-Echo-Drift {coord}: erwartet SOLL {soll_i} € (exakte Zahl), live={val!r}"
            )

    return None


def validate_file(path: Path) -> int:
    resolved = _resolve_profil(path)
    if resolved is None:
        return _fail(
            path,
            "UNKNOWN",
            "kein bekanntes Profil (Spec-Scope = Rebalancing_Tool / "
            "Satelliten_Monitor / Watchlist_Ersatzbank_Monitor) — fail-close",
        )
    profil, prof = resolved

    # Punkt A: Open-Repair (Exception/Load-Fail = strukturelle Korruption).
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except (InvalidFileException, KeyError, OSError, ValueError) as e:
        return _fail(path, profil, f"openpyxl load failed (Punkt A): {e}")

    try:
        # Punkt A (Sheet-Existenz): Workbook geladen (oben) + hat >=1 Worksheet.
        # Kanonische §18.7-Doktrin (xlsx-smoke-test.md Punkt A) = "Datei lesbar
        # + erwartete Sheets vorhanden" — KEIN A1-non-empty-Constraint fürs
        # Voll-Profil (Satelliten_Monitor hat legitim A1=None; Header/Daten
        # beginnen tiefer). Der A1-Check bleibt Minimal-Profil-Proxy (siehe
        # unten) — dort ist er die designierte Akzeptanz-#2-bad-Fixture.
        if not wb.worksheets:
            return _fail(path, profil, "Workbook hat keine Worksheets (Punkt A)")

        for sheet in prof["sheets"]:
            if sheet not in wb.sheetnames:
                return _fail(path, profil, f"erwartetes Sheet '{sheet}' fehlt")

        if prof["scope"] == "minimal":
            # Watchlist-Minimal-Annex: A1 als konkreter Lesbarkeits-/Nicht-
            # Degeneriert-Proxy. Akzeptanz-#2-bad-Fixture = empty_a1 → FAIL,
            # clean = A1 gesetzt → PASS (_smoke_test.py + _generate_fixtures.py).
            if wb.worksheets[0]["A1"].value in (None, ""):
                return _fail(path, profil, "A1 leer — Existenz-Proxy (Punkt A, minimal)")
            return 0  # Watchlist: nur A1 + Existenz (0 Formeln/CF)

        # Punkt B: Error-Token-Scan über alle Zellen aller Sheets.
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and any(t in val for t in _ERROR_TOKENS):
                        return _fail(
                            path,
                            profil,
                            f"Formel-/Wert-Fehler in {ws.title}!{cell.coordinate}: "
                            f"{val!r} (Punkt B)",
                        )

        # Punkt E: Conditional-Format-Regel-Count.
        cf_count = _count_cf_rules(wb)
        if cf_count != prof["cf_rule_count"]:
            return _fail(
                path,
                profil,
                f"Conditional-Format-Count {cf_count} != erwartet "
                f"{prof['cf_rule_count']} — CF von openpyxl-Write zerstört? "
                f"(Punkt E; Soll-SSoT = xlsx-smoke-test.md)",
            )

        # Punkt G: Sparrate-Σ-Sanity (nur Satelliten_Monitor-Profil, SPEC §4.3).
        if profil == "Satelliten_Monitor":
            g_err = _check_g_sparrate_sigma(wb, _CONFIG_YAML_PATH)
            if g_err is not None:
                return _fail(path, profil, f"Sparrate-Σ-Drift (Punkt G): {g_err}")

        return 0
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    paths = list(sys.argv[1:]) if argv is None else argv
    for p in paths:
        path = Path(p)
        if not path.is_file():
            continue
        rc = validate_file(path)
        if rc != 0:
            return rc
    return 0


if __name__ == "__main__":
    sys.exit(main())
